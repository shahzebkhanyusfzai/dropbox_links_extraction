import dropbox
import openpyxl



def extract_links_from_xlsx(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    
    # Get the active sheet
    sheet = workbook.active
    
    # Initialize an empty list to store the links
    links_list = []
    row_num = 1
    # Loop through column H and extract links
    for row in sheet.iter_rows(min_col=8, max_col=8, min_row=1, values_only=True):
        link = row[0]
        if link and "Dropbox" in link:
            # Get the part after "Dropbox"
            manipulated_link = link.split("Dropbox", 1)[1]
            
            # Convert backslashes to forward slashes for Python-friendly paths
            manipulated_link = manipulated_link.replace("\\", "/")
            
            # Ensure the link starts with a "/"
            if not manipulated_link.startswith("/"):
                manipulated_link = "/" + manipulated_link
            
            # Add the manipulated link to the list
            links_list.append((manipulated_link, row_num))
        row_num += 1
            
    return workbook, sheet, links_list




def create_shared_link(dbx, file_path):
    """Create a shared link for the given file_path using the Dropbox SDK."""
    try:
        # Create or get a shared link for the file
        shared_link_metadata = dbx.sharing_create_shared_link(file_path)
        
        # Return the URL of the shared link
        return shared_link_metadata.url
        
    except dropbox.exceptions.ApiError as e:
        # print(f"Failed to create shared link for {file_path}. Reason: {e}")
        return None

# Authenticate with Dropbox
# ACCESS_TOKEN = "yi9cjtevnvycgs2"
# ACCESS_TOKEN = "sl.BmYhk5C7msd2sZkyRsdnudbYKDmqEzc8-sd-tyAgax136e7sisdiyxwogN5p-axv32k"
ACCESS_TOKEN = "sl.Bms83umw8IMnUTCtsdtiFb4u4fZzYiy3Bk7sqvmEGP-1CPp_RAMAM-kGBuHq7PDNfULjM"
dbx = dropbox.Dropbox(ACCESS_TOKEN)

# Specify the file path in Dropbox for which you want to create a shared link
# This should be the full path as it appears in Dropbox, including the starting '/'
# For example: '/folder_name/file_name.zip'
xlsm_file_path = "C:/Users/shahz/Desktop/Test-Dropbox-NEW.xlsx"
# xlsm_file_path = "D:/Dropbox/_ECOMMERCE/ETSY/Listing/Dropbox-files.xlsx"
# dropbox_file_path = "/sha/shah.zip"
workbook, sheet, dropbox_file_path_list = extract_links_from_xlsx(xlsm_file_path)

# Create the shared link and capture the URL
for dropbox_link, row_num in dropbox_file_path_list:
    print(dropbox_link)
    shared_link_url = create_shared_link(dbx, dropbox_link)
    if shared_link_url:
        print(f"Shared link for {dropbox_link}: {shared_link_url}")
        sheet.cell(row=row_num, column=19).value = shared_link_url  # column 9 corresponds to column I


    else:
        print(f"Could not generate shared link for {dropbox_link}.")

workbook.save(xlsm_file_path)
